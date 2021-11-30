import openpyxl
from copy import copy
wb = openpyxl.load_workbook("1.xlsx")
ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]
def copy_style():
    for row in ws1:
        for cell in row:
            if cell.has_style:
                ws2[cell.coordinate].font = copy(cell.font)
                # ws2[cell.coordinate].font = cell.font
                ws2[cell.coordinate].border = copy(cell.border)
                ws2[cell.coordinate].fill = copy(cell.fill)
                # 数字格式复制
                ws2[cell.coordinate].number_format = copy(cell.number_format)
                ws2[cell.coordinate].protection = copy(cell.protection)
                ws2[cell.coordinate].alignment = copy(cell.alignment)
                # print(cell.font)
                # print(cell.border)
                # print(cell.number_format)
                # print(cell.protection)
# print(ws)

def process_merged(worksheet_1, worksheet_2):
    for range in worksheet_1.merged_cells:
        worksheet_2.merge_cells(str(range))
        # print(str(range))

def copy_value(worksheet_1, worksheet_2):
    for row in worksheet_1:
        for cell in row:
            worksheet_2[cell.coordinate].value = cell.value

def unmerge(worksheet):
    while worksheet.merged_cells:
        for m in worksheet.merged_cells:
            print(type(m), str(m))
            worksheet.unmerge_cells(str(m))
        # break
if __name__ == '__main__':
    # copy_style()
    unmerge(ws2)
    copy_value(ws1, ws2)
    process_merged(ws1, ws2)
    wb.save("1.xlsx")
