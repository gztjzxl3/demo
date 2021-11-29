import openpyxl

def csv_to_xlsx(csv_file_name,sheet_name='Sheet'):
    wb = openpyxl.Workbook()
    if sheet_name != 'Sheet':
        sheet = wb.create_sheet(sheet_name)
    else:
        sheet = wb.get_sheet_by_name(sheet_name)
    fp = open(csv_file_name, 'r')
    text = fp.read()
    data = text.split('\n')[:-1]
    for row_n in range(len(data)):
        row_list = data[row_n].split(',')
        for col_n in range(len(row_list)):
            sheet.cell(row_n + 1, col_n + 1).value = row_list[col_n]
    wb.save(csv_file_name.replace('.csv','.xlsx'))
    fp.close()
#
if __name__ == '__main__':
    csv_to_xlsx('data_2.csv')
