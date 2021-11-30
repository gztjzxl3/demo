import openpyxl
import time

# wb = openpyxl.Workbook()

# ws.merge_cells(start_row=1,start_column=1,end_row=10,end_column=3)
# ws.unmerge_cells(start_row=1,start_column=1,end_row=10,end_column=20)
# sheet = wb.get_sheet_by_name('Sheet')
# wb.create_sheet('data')
# sheet_now = wb.active
# print(sheet_now, type(sheet_now), sheet_now.title)
# print(wb.get_sheet_names(), '\n', type(wb.get_sheet_names))
# sheet = wb.active
# sheet['A1'] = 12
# print(sheet_now.title)
# print(ws.merged_cells)
# ws.unmerge_cells('A1:C3')
def unmerge(worksheet):
    while worksheet.merged_cells:
        for m in worksheet.merged_cells:
            print(type(m), str(m))
            ws.unmerge_cells(str(m))
            break
# for ws in wb:
#     print(ws.title)
if __name__ == '__main__':
    wb = openpyxl.load_workbook('1.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    unmerge(ws)
    wb.save('1.xlsx')
